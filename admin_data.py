"""Admin-safe edits for the local Excel data source."""

from __future__ import annotations

from datetime import datetime, timedelta
from pathlib import Path, PurePosixPath
import re
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

import clean_workbook

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover - dependency checked at runtime
    load_workbook = None

ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "sample_data.xlsx"

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"

TABLES = [
    "Departments",
    "Employees",
    "Projects",
    "Tasks",
    "Meetings",
    "Weekly Updates",
    "Activity Log",
    "Lists",
]
KEY_FIELDS = {
    "Departments": "Department ID",
    "Employees": "Employee ID",
    "Projects": "Project ID",
    "Tasks": "Task ID",
    "Meetings": "Meeting ID",
    "Weekly Updates": "Update ID",
    "Activity Log": "Activity ID",
    "Lists": "Statuses",
}
DATE_FIELDS = {
    "Employees": {"Hire Date"},
    "Projects": {"Start Date", "Target End Date"},
    "Tasks": {"Due Date"},
    "Meetings": {"Date/Time"},
    "Weekly Updates": {"Week Starting"},
    "Activity Log": {"Timestamp"},
}
NUMBER_FIELDS = {
    "Departments": {"Headcount", "Annual Budget SAR"},
    "Projects": {"Progress %", "Budget SAR", "Actual Spend SAR"},
    "Tasks": {"Estimated Hours", "Actual Hours", "Completion %"},
    "Meetings": {"Duration Minutes", "Attendees Count"},
    "Weekly Updates": {"Progress %"},
}
REQUIRED_FIELDS = {
    "Departments": {"Department ID", "Department Name", "Division", "Location"},
    "Employees": {
        "Employee ID",
        "Employee Name",
        "Email",
        "Department ID",
        "Department",
        "Job Title",
        "Level",
        "Hire Date",
        "Employment Status",
    },
    "Projects": {
        "Project ID",
        "Project Name",
        "Department ID",
        "Department",
        "Owner ID",
        "Owner",
        "Status",
        "Priority",
        "Risk Level",
        "Start Date",
        "Target End Date",
        "Progress %",
    },
    "Tasks": {
        "Task ID",
        "Project ID",
        "Project",
        "Task Name",
        "Assigned To ID",
        "Assigned To",
        "Department ID",
        "Department",
        "Status",
        "Priority",
        "Due Date",
        "Completion %",
    },
    "Meetings": {
        "Meeting ID",
        "Project ID",
        "Project",
        "Meeting Type",
        "Date/Time",
        "Organizer ID",
        "Organizer",
        "Outcome",
    },
    "Weekly Updates": {
        "Update ID",
        "Week Starting",
        "Project ID",
        "Project",
        "Department",
        "Health",
        "Status",
        "Progress %",
        "Next Step",
    },
    "Activity Log": {
        "Activity ID",
        "Timestamp",
        "Employee ID",
        "Employee",
        "Department ID",
        "Department",
        "Project ID",
        "Project",
        "Activity Type",
        "Impact",
        "Source",
    },
    "Lists": {"Statuses"},
}


class AdminDataError(ValueError):
    """Safe validation failure for admin data edits."""


def qname(name: str) -> str:
    return f"{{{MAIN_NS}}}{name}"


def excel_date(value: int | float) -> str:
    moment = datetime(1899, 12, 30) + timedelta(days=float(value))
    if moment.time() == datetime.min.time():
        return moment.date().isoformat()
    return moment.isoformat(timespec="minutes")


def date_serial(value: str) -> float:
    if not value:
        raise AdminDataError("Date fields cannot be blank.")
    cleaned = value.strip()
    try:
        moment = datetime.fromisoformat(cleaned)
    except ValueError as error:
        raise AdminDataError(f"Invalid date value: {value}") from error
    base = datetime(1899, 12, 30)
    delta = moment - base
    serial = delta.days + delta.seconds / 86400
    return int(serial) if serial.is_integer() else round(serial, 10)


def column_name(index: int) -> str:
    name = ""
    index += 1
    while index:
        index, remainder = divmod(index - 1, 26)
        name = chr(65 + remainder) + name
    return name


def sheet_paths(archive: ZipFile) -> dict[str, str]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    targets = {node.attrib["Id"]: node.attrib["Target"] for node in relationships}
    sheets = workbook.find(qname("sheets"))
    if sheets is None:
        raise AdminDataError("Workbook contains no worksheets.")
    return {
        sheet.attrib["name"]: (
            targets[sheet.attrib[f"{{{REL_NS}}}id"]].lstrip("/")
            if targets[sheet.attrib[f"{{{REL_NS}}}id"]].startswith("/xl/")
            else str(PurePosixPath("xl") / targets[sheet.attrib[f"{{{REL_NS}}}id"]])
        )
        for sheet in sheets
    }


def shared_strings(archive: ZipFile) -> list[str]:
    root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
    return [
        "".join(node.text or "" for node in item.iter(qname("t")))
        for item in root.findall(qname("si"))
    ]


def cell_value(cell: ET.Element, strings: list[str]):
    value = cell.find(qname("v"))
    inline = cell.find(qname("is"))
    if inline is not None:
        return "".join(node.text or "" for node in inline.iter(qname("t")))
    if value is None:
        return None
    raw = value.text or ""
    if cell.attrib.get("t") == "s":
        return strings[int(raw)]
    try:
        return float(raw) if "." in raw else int(raw)
    except ValueError:
        return raw


def column_index(reference: str) -> int:
    letters = re.match(r"[A-Z]+", reference)
    if not letters:
        raise AdminDataError(f"Invalid cell reference: {reference}")
    result = 0
    for letter in letters.group():
        result = result * 26 + ord(letter) - 64
    return result - 1


def read_table(table: str) -> tuple[list[str], list[dict[str, object]]]:
    if table not in TABLES:
        raise AdminDataError("Unknown table.")
    if load_workbook is not None:
        workbook = load_workbook(SOURCE, data_only=True)
        worksheet = workbook[table]
        headers = [str(cell.value or "") for cell in worksheet[1]]
        records = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            record = dict(zip(headers, row))
            if not record.get(headers[0]):
                continue
            for field in DATE_FIELDS.get(table, set()):
                value = record.get(field)
                if isinstance(value, datetime):
                    record[field] = (
                        value.date().isoformat()
                        if value.time() == datetime.min.time()
                        else value.isoformat(timespec="minutes")
                    )
            records.append(record)
        return headers, records
    with ZipFile(SOURCE) as archive:
        paths = sheet_paths(archive)
        strings = shared_strings(archive)
        root = ET.fromstring(archive.read(paths[table]))
        rows = []
        for row_node in root.findall(f".//{qname('sheetData')}/{qname('row')}"):
            cells: dict[int, object] = {}
            for cell in row_node.findall(qname("c")):
                cells[column_index(cell.attrib["r"])] = cell_value(cell, strings)
            if cells:
                row = [None] * (max(cells) + 1)
                for index, value in cells.items():
                    row[index] = value
                rows.append(row)
    if not rows:
        raise AdminDataError("Table has no header row.")
    headers = [str(value or "") for value in rows[0]]
    records = []
    for row in rows[1:]:
        padded = row + [None] * (len(headers) - len(row))
        record = dict(zip(headers, padded))
        if not record.get(headers[0]):
            continue
        for field in DATE_FIELDS.get(table, set()):
            if record.get(field) is not None:
                record[field] = excel_date(record[field])
        records.append(record)
    return headers, records


def normalize_record(table: str, headers: list[str], record: dict[str, object]) -> dict[str, object]:
    normalized = {}
    for header in headers:
        value = record.get(header)
        if value == "":
            value = None
        if header in NUMBER_FIELDS.get(table, set()) and value is not None:
            try:
                number = float(value)
            except (TypeError, ValueError) as error:
                raise AdminDataError(f"{header} must be a number.") from error
            value = int(number) if number.is_integer() else number
        normalized[header] = value
    key = KEY_FIELDS[table]
    if not normalized.get(key):
        raise AdminDataError(f"{key} is required.")
    missing = [
        field
        for field in REQUIRED_FIELDS.get(table, {key})
        if field in headers and normalized.get(field) in (None, "")
    ]
    if missing:
        raise AdminDataError(f"Please fill the required fields: {', '.join(missing)}.")
    return normalized


def next_employee_id(records: list[dict[str, object]]) -> str:
    highest = 0
    for record in records:
        value = str(record.get("Employee ID") or "")
        match = re.fullmatch(r"E(\d+)", value)
        if match:
            highest = max(highest, int(match.group(1)))
    return f"E{highest + 1:04d}"


def cell_xml(reference: str, value: object, is_date: bool = False) -> str:
    if value is None:
        return ""
    if is_date:
        value = date_serial(str(value))
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return f'<c r="{reference}"><v>{value}</v></c>'
    text = escape(str(value), {'"': "&quot;"})
    return f'<c r="{reference}" t="inlineStr"><is><t>{text}</t></is></c>'


def build_sheet_data(table: str, headers: list[str], records: list[dict[str, object]]) -> str:
    rows = []
    header_cells = "".join(
        cell_xml(f"{column_name(index)}1", header)
        for index, header in enumerate(headers)
    )
    rows.append(f'<row r="1">{header_cells}</row>')
    date_fields = DATE_FIELDS.get(table, set())
    for row_index, record in enumerate(records, start=2):
        cells = "".join(
            cell_xml(
                f"{column_name(column)}{row_index}",
                record.get(header),
                header in date_fields,
            )
            for column, header in enumerate(headers)
        )
        rows.append(f'<row r="{row_index}">{cells}</row>')
    return f"<sheetData>{''.join(rows)}</sheetData>"


def replace_sheet_data(xml: bytes, table: str, headers: list[str], records: list[dict[str, object]]) -> bytes:
    text = xml.decode("utf-8")
    last_column = column_name(max(len(headers) - 1, 0))
    last_row = max(len(records) + 1, 1)
    dimension = f'A1:{last_column}{last_row}'
    text = re.sub(r'<dimension ref="[^"]+"\s*/>', f'<dimension ref="{dimension}"/>', text, count=1)
    replacement = build_sheet_data(table, headers, records)
    updated, count = re.subn(r"<sheetData>.*?</sheetData>", replacement, text, count=1, flags=re.S)
    if count != 1:
        raise AdminDataError("Could not update worksheet data.")
    return updated.encode("utf-8")


def write_table(table: str, headers: list[str], records: list[dict[str, object]]) -> None:
    if load_workbook is not None:
        workbook = load_workbook(SOURCE)
        worksheet = workbook[table]
        if worksheet.max_row > 1:
            worksheet.delete_rows(2, worksheet.max_row - 1)
        date_fields = DATE_FIELDS.get(table, set())
        for row_index, record in enumerate(records, start=2):
            for column_index, header in enumerate(headers, start=1):
                value = record.get(header)
                if header in date_fields and value:
                    value = datetime.fromisoformat(str(value))
                worksheet.cell(row_index, column_index).value = value
        workbook.save(SOURCE)
        clean_workbook.clean()
        return
    temp_path = SOURCE.with_suffix(".admin.tmp.xlsx")
    with ZipFile(SOURCE) as source:
        paths = sheet_paths(source)
        sheet_path = paths[table]
        sheet_xml = replace_sheet_data(source.read(sheet_path), table, headers, records)
        with ZipFile(temp_path, "w", ZIP_DEFLATED) as target:
            for info in source.infolist():
                target.writestr(info, sheet_xml if info.filename == sheet_path else source.read(info))
    temp_path.replace(SOURCE)
    clean_workbook.clean()


def restore_table_records(table: str, records: list[dict[str, object]]) -> dict[str, object]:
    if table not in TABLES:
        raise AdminDataError("Unknown table.")
    headers, _existing = read_table(table)
    normalized_records = [
        normalize_record(table, headers, record if isinstance(record, dict) else {})
        for record in records
    ]
    write_table(table, headers, normalized_records)
    return {"table": table, "action": "restore", "count": len(normalized_records)}


def apply_record_action(table: str, action: str, record: dict[str, object]) -> dict[str, object]:
    if action not in {"add", "update", "delete"}:
        raise AdminDataError("Unknown admin action.")
    headers, records = read_table(table)
    key = KEY_FIELDS[table]
    if table == "Employees" and action == "add" and not str(record.get(key) or "").strip():
        record = {**record, key: next_employee_id(records)}
    normalized = normalize_record(table, headers, record)
    key_value = str(normalized[key])
    existing_index = next(
        (index for index, item in enumerate(records) if str(item.get(key)) == key_value),
        None,
    )
    if action == "add":
        if existing_index is not None:
            raise AdminDataError(f"{key} already exists.")
        records.append(normalized)
    elif action == "update":
        if existing_index is None:
            raise AdminDataError(f"{key} was not found.")
        records[existing_index] = normalized
    else:
        if existing_index is None:
            raise AdminDataError(f"{key} was not found.")
        records.pop(existing_index)
    write_table(table, headers, records)
    return {"table": table, "action": action, "key": key, "keyValue": key_value}
