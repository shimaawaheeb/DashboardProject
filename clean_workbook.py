#!/usr/bin/env python3
"""Create a cleaned workbook without reserializing Excel's XML documents."""

from __future__ import annotations

from pathlib import Path, PurePosixPath
import re
from xml.etree import ElementTree as ET
from xml.sax.saxutils import escape
from zipfile import ZIP_DEFLATED, ZipFile

from config import CLEANED_WORKBOOK, SAMPLE_WORKBOOK, TEMP_CLEANED_WORKBOOK

ROOT = Path(__file__).resolve().parent
SOURCE = SAMPLE_WORKBOOK
OUTPUT = CLEANED_WORKBOOK
TEMP_OUTPUT = TEMP_CLEANED_WORKBOOK

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"


def qname(name: str) -> str:
    return f"{{{MAIN_NS}}}{name}"


def read_shared_strings(root: ET.Element) -> list[str]:
    return [
        "".join(node.text or "" for node in item.iter(qname("t")))
        for item in root.findall(qname("si"))
    ]


def cell_text(cell: ET.Element, shared_strings: list[str]) -> str | None:
    value = cell.find(qname("v"))
    if value is None:
        return None
    if cell.attrib.get("t") == "s":
        return shared_strings[int(value.text)]
    return value.text


def sheet_paths(archive: ZipFile) -> dict[str, str]:
    workbook = ET.fromstring(archive.read("xl/workbook.xml"))
    relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
    targets = {node.attrib["Id"]: node.attrib["Target"] for node in relationships}
    sheets = workbook.find(qname("sheets"))
    if sheets is None:
        raise ValueError("Workbook contains no worksheets")
    return {
        sheet.attrib["name"]: (
            targets[sheet.attrib[f"{{{REL_NS}}}id"]].lstrip("/")
            if targets[sheet.attrib[f"{{{REL_NS}}}id"]].startswith("/xl/")
            else str(PurePosixPath("xl") / targets[sheet.attrib[f"{{{REL_NS}}}id"]])
        )
        for sheet in sheets
    }


def sheet_records(
    xml: bytes, shared_strings: list[str]
) -> list[tuple[dict[str, str | None], dict[str, str]]]:
    """Read row values and map column headers to cell references."""
    root = ET.fromstring(xml)
    rows = root.findall(f".//{qname('sheetData')}/{qname('row')}")
    if not rows:
        return []

    header_cells = rows[0].findall(qname("c"))
    columns = {
        "".join(filter(str.isalpha, cell.attrib["r"])): cell_text(
            cell, shared_strings
        )
        for cell in header_cells
    }
    records = []
    for row in rows[1:]:
        values: dict[str, str | None] = {}
        references: dict[str, str] = {}
        for cell in row.findall(qname("c")):
            reference = cell.attrib["r"]
            column = "".join(filter(str.isalpha, reference))
            header = columns.get(column)
            if header:
                values[header] = cell_text(cell, shared_strings)
                references[header] = reference
        if values:
            records.append((values, references))
    return records


def replace_shared_string_index(xml: bytes, reference: str, index: int) -> bytes:
    """Change one shared-string cell index while preserving all other XML bytes."""
    text = xml.decode("utf-8")
    pattern = re.compile(
        rf'(<c\b(?=[^>]*\br="{re.escape(reference)}")[^>]*'
        rf'\bt="s"[^>]*>.*?<v>)(\d+)(</v>)'
    )
    updated, count = pattern.subn(
        lambda match: f"{match.group(1)}{index}{match.group(3)}",
        text,
        count=1,
    )
    if count != 1:
        raise ValueError(f"Could not update shared-string cell {reference}")
    return updated.encode("utf-8")


def append_shared_strings(xml: bytes, values: list[str]) -> bytes:
    """Append plain shared strings without changing Excel namespace declarations."""
    if not values:
        return xml
    text = xml.decode("utf-8")
    root_match = re.search(r"<sst\b[^>]*>", text)
    if not root_match:
        raise ValueError("Invalid sharedStrings.xml root")

    root_tag = root_match.group(0)
    unique_match = re.search(r'\buniqueCount="(\d+)"', root_tag)
    if not unique_match:
        raise ValueError("sharedStrings.xml has no uniqueCount")
    unique_count = int(unique_match.group(1)) + len(values)
    updated_root = re.sub(
        r'\buniqueCount="\d+"',
        f'uniqueCount="{unique_count}"',
        root_tag,
        count=1,
    )
    text = text[: root_match.start()] + updated_root + text[root_match.end() :]

    additions = "".join(f"<si><t>{escape(value)}</t></si>" for value in values)
    closing = text.rfind("</sst>")
    if closing == -1:
        raise ValueError("Invalid sharedStrings.xml closing tag")
    return (text[:closing] + additions + text[closing:]).encode("utf-8")


def clean() -> dict[str, int]:
    with ZipFile(SOURCE) as source:
        if "xl/sharedStrings.xml" not in source.namelist():
            return clean_with_openpyxl()
        paths = sheet_paths(source)
        shared_xml = source.read("xl/sharedStrings.xml")
        shared_strings = read_shared_strings(ET.fromstring(shared_xml))
        indexes = {value: index for index, value in enumerate(shared_strings)}
        appended_strings: list[str] = []
        replacements: dict[str, bytes] = {}

        employees_path = paths["Employees"]
        employees_xml = source.read(employees_path)
        seen_emails: set[str] = set()
        email_changes = 0
        for values, references in sheet_records(employees_xml, shared_strings):
            email = (values.get("Email") or "").strip().lower()
            if email and "@" in email and email in seen_emails:
                local, domain = email.split("@", 1)
                employee_id = (values.get("Employee ID") or "").lower()
                cleaned_email = f"{local}.{employee_id}@{domain}"
                index = indexes.get(cleaned_email)
                if index is None:
                    index = len(shared_strings) + len(appended_strings)
                    indexes[cleaned_email] = index
                    appended_strings.append(cleaned_email)
                employees_xml = replace_shared_string_index(
                    employees_xml, references["Email"], index
                )
                email_changes += 1
            if email:
                seen_emails.add(email)
        replacements[employees_path] = employees_xml

        tasks_path = paths["Tasks"]
        tasks_xml = source.read(tasks_path)
        completed_index = indexes["Completed"]
        status_changes = 0
        for values, references in sheet_records(tasks_xml, shared_strings):
            if (
                values.get("Completion %") == "100"
                and values.get("Status") != "Completed"
            ):
                tasks_xml = replace_shared_string_index(
                    tasks_xml, references["Status"], completed_index
                )
                status_changes += 1
        replacements[tasks_path] = tasks_xml
        replacements["xl/sharedStrings.xml"] = append_shared_strings(
            shared_xml, appended_strings
        )

        with ZipFile(TEMP_OUTPUT, "w", ZIP_DEFLATED) as target:
            for info in source.infolist():
                target.writestr(
                    info, replacements.get(info.filename, source.read(info))
                )

    TEMP_OUTPUT.replace(OUTPUT)
    return {
        "duplicate_emails_fixed": email_changes,
        "task_statuses_fixed": status_changes,
    }


def clean_with_openpyxl() -> dict[str, int]:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise RuntimeError(
            "This workbook uses inline strings. Install openpyxl to clean it."
        ) from error

    workbook = load_workbook(SOURCE)
    email_changes = 0
    status_changes = 0

    employees = workbook["Employees"]
    headers = {cell.value: index for index, cell in enumerate(employees[1], start=1)}
    seen_emails: set[str] = set()
    for row in range(2, employees.max_row + 1):
        email_cell = employees.cell(row, headers["Email"])
        email = str(email_cell.value or "").strip().lower()
        if email and "@" in email and email in seen_emails:
            local, domain = email.split("@", 1)
            employee_id = str(employees.cell(row, headers["Employee ID"]).value or "").lower()
            email_cell.value = f"{local}.{employee_id}@{domain}"
            email_changes += 1
        if email:
            seen_emails.add(email)

    tasks = workbook["Tasks"]
    headers = {cell.value: index for index, cell in enumerate(tasks[1], start=1)}
    for row in range(2, tasks.max_row + 1):
        completion = tasks.cell(row, headers["Completion %"]).value
        status_cell = tasks.cell(row, headers["Status"])
        if completion == 100 and status_cell.value != "Completed":
            status_cell.value = "Completed"
            status_changes += 1

    workbook.save(OUTPUT)
    return {
        "duplicate_emails_fixed": email_changes,
        "task_statuses_fixed": status_changes,
    }


if __name__ == "__main__":
    results = clean()
    print(f"Created {OUTPUT.name}")
    for check, count in results.items():
        print(f"{check}: {count}")
